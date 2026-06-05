"""
tests/test_excel_roundtrip.py

Verifies that the schema → Excel → schema round trip is lossless at the
top-level slot layer.  The test:

  1. Runs schema_to_excel to generate a workbook in a temp directory.
  2. Runs the comparison logic from excel_to_schema against that workbook.
  3. Asserts that there are zero differences for every main data class.

This test is intentionally self-contained: it imports the helper functions
directly rather than subprocess-calling the scripts, so failures produce
Python tracebacks that point to the exact line.
"""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

# ── path setup ────────────────────────────────────────────────────────────────
_ROOT    = Path(__file__).parent.parent
_SCRIPTS = _ROOT / "scripts"
_SCHEMA  = _ROOT / "src" / "coremeta4cat" / "schema"

sys.path.insert(0, str(_SCRIPTS))

from generate_schema_docs import (   # noqa: E402
    get_all_class_slots,
    get_class_ranged_slot_usage,
    get_slot_details,
    load_merged_schema,
    snake_to_readable,
)
from schema_to_excel import (        # noqa: E402
    CLASS_MAP,
    build_catcore_sheet,
    build_intro_sheet,
    build_legend_sheet,
    build_sheet,
)

# ── constants ─────────────────────────────────────────────────────────────────

# Sheet names are the keys of CLASS_MAP, in a stable order
SHEET_NAMES = ["Synthesis", "Characterization", "Reaction", "Simulation"]


# ── fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def schema():
    """Load the merged schema once for all tests in this module."""
    return load_merged_schema(str(_SCHEMA))


@pytest.fixture(scope="module")
def generated_workbook(schema, tmp_path_factory):
    """Generate the vocabulary workbook into a temporary file."""
    tmp = tmp_path_factory.mktemp("excel")
    out_path = tmp / "coremeta4cat_vocabulary.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_intro_sheet(wb)
    build_legend_sheet(wb)
    build_catcore_sheet(wb, schema)
    for sheet_title in SHEET_NAMES:
        build_sheet(wb, schema, sheet_title)

    wb.save(str(out_path))
    return out_path


# ── helper: mirrors _slot_mro from excel_to_schema / schema_to_excel ─────────

def _slot_mro(schema: dict, class_name: str, slot_name: str) -> str:
    class_def = schema.get("classes", {}).get(class_name, {})
    usage = (class_def.get("slot_usage") or {}).get(slot_name) or {}
    slot_def = get_slot_details(schema, slot_name)
    required    = usage.get("required",    slot_def.get("required",    False))
    recommended = usage.get("recommended", slot_def.get("recommended", False))
    if required:
        return "M"
    if recommended:
        return "R"
    return "O"


def _schema_labels_for_class(schema: dict, class_name: str) -> dict[str, str]:
    result: dict[str, str] = {}
    direct_slots = get_all_class_slots(schema, class_name)
    for slot_name in direct_slots:
        label = snake_to_readable(slot_name)
        result[label] = _slot_mro(schema, class_name, slot_name)
    for su_name, _ in get_class_ranged_slot_usage(schema, class_name):
        if su_name not in direct_slots:
            label = snake_to_readable(su_name)
            result[label] = _slot_mro(schema, class_name, su_name)
    return result


def _excel_labels_for_sheet(df: pd.DataFrame, mro_col: str) -> dict[str, str]:
    result: dict[str, str] = {}
    top = df[
        df["domain"].isin(["", "na", float("nan")])
        & (df["type"].str.strip().str.lower() == "slot")
    ]
    for _, row in top.iterrows():
        label = str(row.get("label", "")).strip()
        mro   = str(row.get(mro_col, "")).strip().upper()[:1]
        if label:
            result[label] = mro if mro in ("M", "R", "O") else "O"
    return result


# ── tests ─────────────────────────────────────────────────────────────────────

def test_workbook_has_all_sheets(generated_workbook):
    """All seven expected sheets are present."""
    wb = openpyxl.load_workbook(str(generated_workbook), read_only=True)
    assert set(wb.sheetnames) == {
        "Introduction", "Legend", "CoreMeta4Cat",
        "Synthesis", "Characterization", "Reaction", "Simulation",
    }


def test_data_sheets_have_rows(generated_workbook):
    """Each of the four main data sheets has at least one data row."""
    wb = openpyxl.load_workbook(str(generated_workbook), read_only=True, data_only=True)
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) > 0, f"Sheet '{sheet_name}' has no data rows"


def test_header_columns(generated_workbook):
    """Each data sheet has exactly the expected column headers in order."""
    expected = ["label", "type", "domain", "M / R / O", "range", "uri", "description"]
    wb = openpyxl.load_workbook(str(generated_workbook), read_only=True, data_only=True)
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        header = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        assert header == expected, (
            f"Sheet '{sheet_name}' header mismatch.\n"
            f"  Expected: {expected}\n  Got:      {header}"
        )


def test_type_column_values(generated_workbook):
    """The 'type' column contains only 'slot' or 'class' values."""
    wb = openpyxl.load_workbook(str(generated_workbook), read_only=True, data_only=True)
    for sheet_name in SHEET_NAMES:
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_type = row[1]  # 'type' is column index 1 (0-based)
            assert row_type in ("slot", "class"), (
                f"Sheet '{sheet_name}', row {row_idx}: "
                f"unexpected type value '{row_type}'"
            )


@pytest.mark.parametrize("sheet_title", SHEET_NAMES)
def test_roundtrip_no_slot_differences(schema, generated_workbook, sheet_title):
    """
    The top-level slots written to the workbook match what the schema reports
    for the corresponding class, with correct M/R/O values.

    This is the core round-trip assertion: schema → Excel → compare → no diff.
    """
    schema_class = CLASS_MAP.get(sheet_title, sheet_title)
    schema_labels = _schema_labels_for_class(schema, schema_class)

    xls = pd.ExcelFile(str(generated_workbook))
    df  = xls.parse(sheet_title).fillna("")
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Detect M/R/O column (flexible)
    mro_col = next(
        (c for c in df.columns if "mandatory" in c or "mro" in c
         or c in ("m/r/o", "m / r / o")),
        None,
    )
    assert mro_col is not None, f"Could not find M/R/O column in sheet '{sheet_title}'"

    # Normalise domain column (new name, with legacy fallback)
    if "domain" in df.columns:
        df["domain"] = df["domain"].astype(str).str.strip().str.lower()
    elif "parent" in df.columns:
        df["domain"] = df["parent"].astype(str).str.strip().str.lower()
    else:
        pytest.fail(f"Neither 'domain' nor 'parent' column found in '{sheet_title}'")

    if "type" not in df.columns:
        df["type"] = "slot"

    excel_labels = _excel_labels_for_sheet(df, mro_col)

    schema_set = set(schema_labels)
    excel_set  = set(excel_labels)

    in_schema_not_excel = schema_set - excel_set
    in_excel_not_schema = excel_set - schema_set
    mro_mismatches = {
        label for label in schema_set & excel_set
        if schema_labels[label] != excel_labels[label]
    }

    messages = []
    if in_schema_not_excel:
        msgs = ", ".join(sorted(in_schema_not_excel))
        messages.append(f"In schema but NOT in workbook: {msgs}")
    if in_excel_not_schema:
        msgs = ", ".join(sorted(in_excel_not_schema))
        messages.append(f"In workbook but NOT in schema: {msgs}")
    if mro_mismatches:
        detail = ", ".join(
            f"{l}(schema={schema_labels[l]},xlsx={excel_labels[l]})"
            for l in sorted(mro_mismatches)
        )
        messages.append(f"M/R/O mismatches: {detail}")

    assert not messages, (
        f"Round-trip differences for '{sheet_title}':\n  "
        + "\n  ".join(messages)
    )


def test_class_map_coverage():
    """CLASS_MAP covers exactly the four main sheet titles."""
    assert set(CLASS_MAP.keys()) == set(SHEET_NAMES), (
        f"CLASS_MAP keys {set(CLASS_MAP.keys())} do not match "
        f"SHEET_NAMES {set(SHEET_NAMES)}"
    )


def test_mandatory_slots_present_in_schema(schema):
    """Every slot marked M in the workbook is actually required in the schema."""
    # Re-generate in a temp dir and check against schema
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_intro_sheet(wb)
    build_legend_sheet(wb)
    build_catcore_sheet(wb, schema)
    for sheet_title in SHEET_NAMES:
        build_sheet(wb, schema, sheet_title)

    for sheet_title in SHEET_NAMES:
        ws = wb[sheet_title]
        for row in ws.iter_rows(min_row=2, values_only=True):
            label, row_type, domain, mro = row[0], row[1], row[2], row[3]
            if row_type == "slot" and not domain and mro == "M":
                # Find matching slot name in schema
                schema_class = CLASS_MAP.get(sheet_title, sheet_title)
                slot_name_candidates = [
                    s for s in get_all_class_slots(schema, schema_class)
                    if snake_to_readable(s) == label
                ]
                if slot_name_candidates:
                    slot_name = slot_name_candidates[0]
                    mro_in_schema = _slot_mro(schema, schema_class, slot_name)
                    assert mro_in_schema == "M", (
                        f"Sheet '{sheet_title}', slot '{label}' is M in workbook "
                        f"but {mro_in_schema} in schema"
                    )
